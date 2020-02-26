# This is a canary version of tabulky
from openpyxl import Workbook
from datetime import date
import os
import numpy as np
import pandas as pd
import time
import sys

__location__ = os.path.realpath(
    os.path.join(os.getcwd(), os.path.dirname(__file__)))

main_value_list , first_column_list , second_column_list , third_column_list , results_list , relative , interval = [] , [] , [] , [] , [""] , None , []

if str(input("Chcete načítať dáta zo súboru?\nAno/ Nie\n")).lower() == "ano":
    print("Vytvorte textový súbor z názvom hodnoty a zadajte čísla do jedného riadku oddelené medzerou. Ako desatinná čiarka sa používa znak \".\" (bodka).")
    time.sleep(2)
    if str(input("Ak ste predošlý krok vykonali, zadajte \"1\".")) == "1":
        f = open(os.path.join(__location__, 'hodnoty.TXT'))
        main_value_list = f.read().split(" ")
    else:
        sys.exit()
else:
    #f = open(os.path.join(__location__, 'hodnoty.TXT'))
    #print(f.read())



    while True:
        a = float(input("Hodnoty: "))
        if a != 0:
            main_value_list.append(a)
        else:
            break

main_value_list = [float(i) for i in main_value_list]

number_of_values = len(main_value_list)

for i in range(1,number_of_values+1):
    first_column_list.append(i)

results_list.append(sum(main_value_list)/len(main_value_list))

for i in range(len(main_value_list)):
    second_column_list.append(main_value_list[i]-results_list[1])

results_list.append(0)

for i in range(len(second_column_list)):
    third_column_list.append(abs(second_column_list[i]))

results_list.append(sum(third_column_list)/len(second_column_list))
relative = (results_list[3]/results_list[1])*100
interval.append(results_list[1]-results_list[3]); interval.append(results_list[1]+results_list[3]); first_column_list.append(results_list[0]); main_value_list.append(results_list[1]) ;second_column_list.append(results_list[2]); third_column_list.append(results_list[3])
longest_first_column , longest_main , longest_second_column , longest_third_column = 0 , 0 , 0 , 0

for i in first_column_list:
    if len(str(i)) > longest_first_column: longest_first_column = len(str(i))

for i in main_value_list:
    if len(str(i)) > longest_main: longest_main = len(str(i))

for i in second_column_list:
    if len(str(i)) > longest_second_column: longest_second_column = len(str(i))

for i in third_column_list:
    if len(str(i)) > longest_third_column: longest_third_column = len(str(i))

print("+","-"*(longest_first_column+2),"+","-"*(longest_main+2),"+","-"*(longest_second_column+2),"+","-"*(longest_third_column+2),"+" , sep = "" )

for i in range(len(first_column_list)):
    print("|",first_column_list[i]," "*(longest_first_column-len(str(first_column_list[i])))+"|",round(main_value_list[i],2)," "*(longest_main-len(str(round(main_value_list[i],2))))+"|",round(second_column_list[i],2)," "*(longest_second_column-len(str(round(second_column_list[i],2))))+"|",round(third_column_list[i],2)," "*(longest_third_column-len(str(round(third_column_list[i],2))))+"|")
    print("+","-"*(longest_first_column+2),"+","-"*(longest_main+2),"+","-"*(longest_second_column+2),"+","-"*(longest_third_column+2),"+" , sep = "" )

print("Relatívna odchýlka: ",relative,"%",sep="")
print("Najpravdepodobnejšia hodnota je: (",interval[0]," ; ",interval[1],")",sep="")
print(date.today())

print("\nPrajete si zapísať hodnoty do tabulky?")
write_table = str(input("Ano/ Nie\n")).lower()

combined_list = []
combined_list.append(first_column_list)
combined_list.append(main_value_list)
combined_list.append(second_column_list)
combined_list.append(third_column_list)
converted = np.array(combined_list).T.tolist()

interval_fin = "Najpravdepodobnejšia hodnota je: ("+str(interval[0])+" ; "+str(interval[1])+")"
relative_fin = "Relatívna odchýlka: "+str(relative)+"%"
if write_table == "ano":
    wb = Workbook()
    ws = wb.active
    ws.insert_rows(0)
    for row in converted:
        ws.append(row)
    ws.cell(row = 4, column = 7, value = interval_fin)
    ws.cell(row = 2, column = 7, value = relative_fin)
    wb.save(os.path.join(__location__, "Tabulka.xlsx"))    
